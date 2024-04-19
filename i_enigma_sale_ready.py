import requests
from lxml import etree
from openpyxl import load_workbook
from openpyxl.styles.colors import WHITE, RGB
import warnings

# !/usr/bin/env python3
# -*- coding: utf-8 -*-

# FIXING THE ARGB HEX VALUES ERROR
__old_rgb_set__ = RGB.__set__

# Dictionary of markings
enigma_codes = {
    'ufarm': ['Зикрач Армавир', 1, 4.0, 1.02],
    'volg': ['Копач', 1, 4.0, 1.02],
    'ufamsk': ['Виталий Москва', 1.05, 4.0, 1]
}

# XPath expression to extract the currency rate
x_dollar = '//*[@id="content"]/div/div/div/div[3]/div/table/tbody/tr[15]/td[5]'
x_euro = '//*[@id="content"]/div/div/div/div[3]/div/table/tbody/tr[16]/td[5]'


def __rgb_set_fixed__(self, instance, value):
    try:
        __old_rgb_set__(self, instance, value)
    except ValueError as e:
        if e.args[0] == 'Colors must be aRGB hex values':
            __old_rgb_set__(self, instance, WHITE)


def parsing_currency():
    # URL of the currency website
    url = "https://www.cbr.ru/currency_base/daily/"

    # Send a GET request to the website
    response = requests.get(url)

    # Check if the request was successful
    if response.status_code == 200:
        # Parse the HTML content
        html_content = response.text
        parser = etree.HTMLParser()
        tree = etree.fromstring(html_content, parser)

        # Find the currency rate element using XPath
        euro_element = tree.xpath(x_euro)
        dollar_element = tree.xpath(x_dollar)

        # Check if the element was found
        if euro_element and dollar_element:
            # Get the text content of the element
            dollar_rate = dollar_element[0].text
            dollar_rate = dollar_rate.replace(",", ".")
            euro_rate = euro_element[0].text
            euro_rate = euro_rate.replace(",", ".")

            # Print the currency rate
            print(f"USD to RUB currency rate: {euro_rate}")
            return float(euro_rate), float(dollar_rate)
        else:
            print("Currency rate element not found.")
    else:
        print("Failed to fetch the webpage.")
        exit()


def load_invoice(enigma_path):
    try:
        enigma_workbook = load_workbook(enigma_path)
        enigma_worksheet = enigma_workbook.active

    except Exception as e:
        print(f"Error has occured: {e}")
        exit()

    return enigma_workbook, enigma_worksheet


def enigma_upd(enigma_workbook, enigma_worksheet, euro_rate, dollar_rate):
    total_rows = enigma_worksheet.max_row
    code_name = str(enigma_worksheet['B5'].value).lower()
    currency_type = str(enigma_worksheet['G14'].value).lower()
    last_flower = 0
    start = 15
    customer = ''
    customer_code = ''
    flower_sum = [0.0, 0.0, 0.0, 0.0]

    for code, info in enigma_codes.items():
        if code in code_name:
            customer = info[0]
            customer_code = code
            break

    column_name = {
        'B14': '', 'C14': 'ФУЛЛ', 'D14': 'ТИП',
        'E14': 'КОЛ-ВО', 'F14': f'ЦЕНА, {currency_type[-3:].upper()}',
        'G14': f'СУММА ЦВЕТОК {currency_type[-3:].upper()}',
        'H14': 'КУРС', 'I14': 'СУММА ЦВЕТОК, РУБ',
        'J14': 'ТРАНСПОРТ, РУБ', 'K14': 'ИТОГО, РУБ',
        'L14': 'ЦЕНА, РУБ'
    }
    sum_total = ['G', 'I', 'J', 'K']

    for row in range(start, total_rows):
        if (
            enigma_worksheet[f'B{row}'].value is None
            and
            last_flower == 0
        ):
            last_flower = row - 1

            if 'Subtotal' in str(enigma_worksheet[f'D{row + 4}'].value):
                flower_sum[1] = float(enigma_worksheet[f'G{row + 4}'].value)

            else:
                print("Subtotal row wasn't found!")
                exit()

            if 'TOTAL FOT' in str(enigma_worksheet[f'F{row+6}'].value):
                flower_sum[0] = float(enigma_worksheet[f'G{row+6}'].value)

            else:
                print("Total row wasn't found!")
                exit()

    if flower_sum[0] != flower_sum[1]:
        print('Extra cost is missing!')
        exit()

    markup_rub = enigma_codes[customer_code][2]
    markup_perc = enigma_codes[customer_code][3]
    flower_perc = enigma_codes[customer_code][1]

    if 'usd' in currency_type:
        currency_rate_upd = (float(dollar_rate) + markup_rub) * markup_perc
        column_name['H14'] = 'КУРС, USD'

    elif 'eur' in currency_type:
        currency_rate_upd = (float(euro_rate) + markup_rub) * markup_perc
        column_name['H14'] = 'КУРС, EUR'

    truck_cost = float(input('Total logistics cost:\n'))
    truck_ratio = truck_cost/flower_sum[0]

    for row in range(start, last_flower + 1):
        subtotal = float(enigma_worksheet[f'G{row}'].value)
        truck_local = truck_ratio * subtotal
        total_rub = (currency_rate_upd * subtotal) * flower_perc
        amount = float(enigma_worksheet[f'E{row}'].value)
        price = (truck_local + total_rub) / amount

        enigma_worksheet[f'H{row}'] = round(currency_rate_upd, 3)
        enigma_worksheet[f'I{row}'] = round(total_rub, 3)
        enigma_worksheet[f'J{row}'] = round(truck_local, 3)
        enigma_worksheet[f'K{row}'] = round(truck_local + total_rub, 3)
        enigma_worksheet[f'L{row}'] = round(price, 3)

        flower_sum[3] = flower_sum[3] + subtotal

    for cell in sum_total:
        enigma_worksheet[f'{cell}{last_flower + 1}'] = f'=SUM({cell}{start}:{cell}{last_flower})'

    for cell, name in column_name.items():
        enigma_worksheet[cell] = name

    if round(flower_sum[0], 2) == round(flower_sum[3], 2):
        print(f'Total flower sale: {flower_sum[3]}')
        enigma_workbook.save(f'EnigmaI {customer}.xlsx')
    else:
        print(flower_sum[0], '/', flower_sum[3])
        print("Totals didn't match!")
        exit()


def start():
    invoice_name = input('Name of the invoice:\n')
    enigma_path = f'{invoice_name}.xlsx'
    enigma_workbook, enigma_worksheet = load_invoice(enigma_path)
    euro_rate, dollar_rate = parsing_currency()
    enigma_upd(enigma_workbook, enigma_worksheet, euro_rate, dollar_rate)


warnings.filterwarnings('ignore', category=UserWarning, module='openpyxl')
RGB.__set__ = __rgb_set_fixed__
start()
